#!/usr/bin/env python3
"""Route workflow v2: city-aware data prep + map + POI-based risk + RA_v2."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import subprocess
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib import parse, request

import folium
import pandas as pd
import pyproj
from folium.plugins import MarkerCluster
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from osgeo import gdal
from shapely.geometry import LineString, Point, Polygon, box, mapping, shape
from shapely.ops import transform, unary_union
from shapely.prepared import prep

BUFFER_EACH_SIDE_M = 100.0
_GEOCODE_CACHE: Dict[Tuple[float, float], Optional[str]] = {}

RA_COLUMNS = [
    "记录时间",
    "航线名称",
    "航线文件链接",
    "地图HTML链接",
    "覆盖城市",
    "航线最高高度(米)",
    "高度风险分级",
    "关键设施风险点",
    "人群聚集场所风险点",
    "活动障碍物风险点(塔吊)",
    "缓冲区风险点总数",
    "环境风险分级",
    "缓冲区平均人口密度(人/平方公里)",
    "人口密度风险分级",
    "总体风险等级",
    "风险点明细",
]


def slugify(value: str) -> str:
    value = value.strip().replace(" ", "_")
    safe = []
    for ch in value:
        if ch.isalnum() or ch in "-_":
            safe.append(ch)
        else:
            safe.append("-")
    result = "".join(safe).strip("-")
    return result or "city"


def run_cmd(cmd: List[str]) -> None:
    print("RUN:", " ".join(str(c) for c in cmd), flush=True)
    subprocess.run(cmd, check=True)


def resolve_kmls(kml_args: Iterable[str]) -> List[Path]:
    files: List[Path] = []
    for item in kml_args:
        p = Path(item)
        if any(ch in item for ch in ["*", "?", "["]):
            files.extend(sorted(Path(".").glob(item)))
        elif p.is_dir():
            files.extend(sorted(p.glob("*.kml")))
        else:
            files.append(p)

    out: List[Path] = []
    seen = set()
    for f in files:
        af = f.resolve()
        if af not in seen:
            seen.add(af)
            out.append(af)
    return out


def parse_kml_coords(kml_path: Path) -> List[Tuple[float, float, float]]:
    text = kml_path.read_text(encoding="utf-8")
    m = re.search(r"<LineString>.*?<coordinates>(.*?)</coordinates>", text, re.DOTALL)
    if not m:
        m = re.search(r"<coordinates>(.*?)</coordinates>", text, re.DOTALL)
    if not m:
        raise ValueError(f"No coordinates in KML: {kml_path}")

    coords: List[Tuple[float, float, float]] = []
    for c in m.group(1).strip().split():
        parts = c.split(",")
        if len(parts) >= 2:
            lon = float(parts[0])
            lat = float(parts[1])
            alt = float(parts[2]) if len(parts) > 2 and parts[2] else 0.0
            coords.append((lat, lon, alt))
    if len(coords) < 2:
        raise ValueError(f"Insufficient coordinates in KML: {kml_path}")
    return coords


def build_projectors(coords: List[Tuple[float, float, float]]):
    center_lon = sum(c[1] for c in coords) / len(coords)
    zone = int((center_lon + 180) / 6) + 1
    south = sum(c[0] for c in coords) / len(coords) < 0
    epsg = 32700 + zone if south else 32600 + zone
    fwd = pyproj.Transformer.from_crs("EPSG:4326", f"EPSG:{epsg}", always_xy=True).transform
    inv = pyproj.Transformer.from_crs(f"EPSG:{epsg}", "EPSG:4326", always_xy=True).transform
    return fwd, inv


def build_route_buffer(coords: List[Tuple[float, float, float]], each_side_m: float) -> Tuple[LineString, Polygon]:
    line_wgs = LineString([(c[1], c[0]) for c in coords])
    fwd, inv = build_projectors(coords)
    line_proj = transform(fwd, line_wgs)
    buffer_proj = line_proj.buffer(each_side_m)
    buffer_wgs = transform(inv, buffer_proj)
    return line_wgs, buffer_wgs


def reverse_geocode_city(lat: float, lon: float) -> Optional[str]:
    key = (round(lat, 4), round(lon, 4))
    if key in _GEOCODE_CACHE:
        return _GEOCODE_CACHE[key]
    params = parse.urlencode(
        {
            "format": "jsonv2",
            "lat": lat,
            "lon": lon,
            "zoom": 8,
            "addressdetails": 1,
        }
    )
    url = f"https://nominatim.openstreetmap.org/reverse?{params}"
    req = request.Request(url, headers={"User-Agent": "risk-workflow-v2/1.0"})
    try:
        with request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        _GEOCODE_CACHE[key] = None
        return None
    addr = data.get("address", {})
    display_name = str(data.get("display_name", ""))
    if display_name:
        for part in [p.strip() for p in display_name.split(",")]:
            if part.endswith("市") and part not in {"中国"}:
                _GEOCODE_CACHE[key] = part
                return part
    city = str(addr.get("city", "")).strip()
    if city and not city.endswith(("区", "县")):
        _GEOCODE_CACHE[key] = city
        return city

    state_district = str(addr.get("state_district", "")).strip()
    if state_district and state_district.endswith("市"):
        _GEOCODE_CACHE[key] = state_district
        return state_district

    for addr_key in ["city", "county", "town", "district"]:
        val = str(addr.get(addr_key, "")).strip()
        if val:
            _GEOCODE_CACHE[key] = val
            return val
    _GEOCODE_CACHE[key] = None
    return None


def detect_cities_for_route(coords: List[Tuple[float, float, float]], buffer_wgs: Polygon) -> List[str]:
    # Use center + endpoints to balance speed and cross-city coverage.
    sample_points: List[Tuple[float, float]] = []
    sample_points.append((coords[0][0], coords[0][1]))
    sample_points.append((coords[-1][0], coords[-1][1]))
    sample_points.append((sum(c[0] for c in coords) / len(coords), sum(c[1] for c in coords) / len(coords)))

    cities: List[str] = []
    seen = set()
    for lat, lon in sample_points:
        city = reverse_geocode_city(lat, lon)
        if city and city not in seen:
            seen.add(city)
            cities.append(city)
    return cities


def city_cache_dir(root: Path, city: str) -> Path:
    return root / "output" / "city_data_cache" / slugify(city)


def _summary_complete(summary: Dict[str, Any]) -> bool:
    outputs = summary.get("outputs", {})
    need = ["poi", "landuse", "hydro", "population", "transport"]
    return all(k in outputs for k in need)


def ensure_city_data(
    root: Path,
    city: str,
    zoom: str = "8-14",
    bbox_override: Optional[Tuple[float, float, float, float]] = None,
) -> Dict[str, Any]:
    cdir = city_cache_dir(root, city)
    cdir.mkdir(parents=True, exist_ok=True)
    summary_path = cdir / "download_summary.json"

    if summary_path.exists():
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            if _summary_complete(summary):
                print(f"[SKIP] city data exists: {city} -> {cdir}", flush=True)
                return summary
        except Exception:
            pass

    script = root / "skills" / "city-data-downloader" / "scripts" / "download_city_data.py"
    cmd = [
        "python3",
        str(script),
        "--city",
        city,
        "--outdir",
        str(cdir),
        "--zoom",
        zoom,
        "--landuse-keys",
        "landuse",
    ]
    if bbox_override is not None:
        south, north, west, east = bbox_override
        cmd.extend(["--bbox", f"{south},{north},{west},{east}"])
    run_cmd(cmd)
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    return summary


def load_poi_rules(path: Path) -> Dict[str, Any]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    required = ["关键设施风险点", "人群聚集场所风险点", "活动障碍物风险点(塔吊)"]
    for k in required:
        if k not in obj:
            raise ValueError(f"POI rules missing category: {k}")
    return obj


def classify_poi(props: Dict[str, Any], rules: Dict[str, Any]) -> str:
    text = " ".join(
        [
            str(props.get("name", "")),
            str(props.get("poi_type", "")),
            str(props.get("raw_tags", "")),
        ]
    ).lower()
    text_norm = re.sub(r"[^a-z0-9]+", " ", text)

    def has_en_token(token: str) -> bool:
        t = token.replace("_", " ")
        return f" {t} " in f" {text_norm} "

    order = ["活动障碍物风险点(塔吊)", "关键设施风险点", "人群聚集场所风险点"]
    for cat in order:
        conf = rules.get(cat, {})
        en = conf.get("keywords_en", [])
        zh = conf.get("keywords_zh", [])
        if any(has_en_token(str(t).lower()) for t in en) or any(str(t) in text for t in zh):
            return cat
    return ""


def parse_raw_tags(raw: Any) -> Dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return {}


def is_water_landuse_feature(feature: Dict[str, Any]) -> bool:
    props = feature.get("properties", {})
    ltype = str(props.get("landuse_type", "")).lower()
    tags = parse_raw_tags(props.get("raw_tags", {}))

    if ltype.startswith("natural:water") or ltype.startswith("waterway:"):
        return True
    if ltype.startswith("landuse:reservoir") or ltype.startswith("landuse:basin"):
        return True

    nat = str(tags.get("natural", "")).lower()
    waterway = str(tags.get("waterway", "")).lower()
    landuse = str(tags.get("landuse", "")).lower()
    if nat in {"water", "wetland"}:
        return True
    if waterway:
        return True
    if landuse in {"reservoir", "basin"}:
        return True
    return False


def load_geojson_features(path: Path) -> List[Dict[str, Any]]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    return obj.get("features", [])


def calc_population_with_water_zero(
    buffer_wgs: Polygon,
    pop_tif: Path,
    water_union: Optional[Polygon],
) -> float:
    ds = gdal.Open(str(pop_tif))
    if ds is None:
        return 0.0
    gt = ds.GetGeoTransform()
    rb = ds.GetRasterBand(1)
    nodata = rb.GetNoDataValue()

    min_lon, min_lat, max_lon, max_lat = buffer_wgs.bounds

    px_a = (min_lon - gt[0]) / gt[1]
    px_b = (max_lon - gt[0]) / gt[1]
    py_a = (max_lat - gt[3]) / gt[5]
    py_b = (min_lat - gt[3]) / gt[5]

    px_min = math.floor(min(px_a, px_b))
    px_max = math.ceil(max(px_a, px_b))
    py_min = math.floor(min(py_a, py_b))
    py_max = math.ceil(max(py_a, py_b))

    off_x = max(0, px_min)
    off_y = max(0, py_min)
    end_x = min(ds.RasterXSize - 1, px_max)
    end_y = min(ds.RasterYSize - 1, py_max)
    win_x = end_x - off_x + 1
    win_y = end_y - off_y + 1
    if win_x <= 0 or win_y <= 0:
        return 0.0

    data = rb.ReadAsArray(off_x, off_y, win_x, win_y)
    if data is None:
        return 0.0

    prep_buf = prep(buffer_wgs)
    prep_water = prep(water_union) if water_union is not None and not water_union.is_empty else None

    weighted_sum = 0.0
    weighted_area = 0.0

    for y in range(win_y):
        for x in range(win_x):
            val = float(data[y, x])
            if nodata is not None and val == nodata:
                continue

            lon0 = gt[0] + (off_x + x) * gt[1]
            lat0 = gt[3] + (off_y + y) * gt[5]
            lon1 = lon0 + gt[1]
            lat1 = lat0 + gt[5]
            cell = box(min(lon0, lon1), min(lat0, lat1), max(lon0, lon1), max(lat0, lat1))

            if not prep_buf.intersects(cell):
                continue
            inter = cell.intersection(buffer_wgs)
            if inter.is_empty:
                continue
            inter_area = inter.area
            if inter_area <= 0:
                continue

            if prep_water is not None and prep_water.intersects(inter):
                water_area = inter.intersection(water_union).area
                land_area = max(inter_area - water_area, 0.0)
            else:
                land_area = inter_area

            weighted_sum += land_area * val
            weighted_area += inter_area

    if weighted_area <= 0:
        return 0.0
    return weighted_sum / weighted_area


def risk_level_from_height(max_alt: float) -> str:
    if max_alt >= 200:
        return "高"
    if max_alt > 120:
        return "中"
    return "低"


def risk_level_from_env(total_points: int) -> str:
    if total_points > 5:
        return "高"
    if total_points >= 1:
        return "中"
    return "低"


def risk_level_from_population(avg_pop: float) -> str:
    if avg_pop > 20000:
        return "高"
    if avg_pop > 5000:
        return "中"
    return "低"


def max_risk_level(*levels: str) -> str:
    order = {"低": 0, "中": 1, "高": 2}
    return max(levels, key=lambda x: order.get(x, 0))


def build_route_html(
    out_html: Path,
    route_name: str,
    coords: List[Tuple[float, float, float]],
    buffer_wgs: Polygon,
    poi_features: List[Dict[str, Any]],
    landuse_features: List[Dict[str, Any]],
    water_surface_features: List[Dict[str, Any]],
    waterway_features: List[Dict[str, Any]],
    population_tile_dirs: List[Tuple[str, Path]],
    road_features: List[Dict[str, Any]],
    hsr_features: List[Dict[str, Any]],
) -> None:
    center_lat = sum(c[0] for c in coords) / len(coords)
    center_lon = sum(c[1] for c in coords) / len(coords)
    route_line = [(c[0], c[1]) for c in coords]

    m = folium.Map(location=[center_lat, center_lon], zoom_start=13, tiles=None, control_scale=True)

    # Prefer high-resolution CN-friendly base tiles first.
    folium.TileLayer(
        tiles="https://webrd0{s}.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=2&style=8&x={x}&y={y}&z={z}",
        attr="高德地图",
        name="高德矢量(高清)",
        overlay=False,
        control=True,
        subdomains="1234",
        max_zoom=20,
        max_native_zoom=20,
        detect_retina=True,
    ).add_to(m)
    folium.TileLayer(
        "OpenStreetMap",
        name="OpenStreetMap",
        overlay=False,
        control=True,
        max_zoom=20,
        max_native_zoom=20,
        detect_retina=True,
    ).add_to(m)
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri World Imagery",
        name="卫星图",
        overlay=False,
        control=True,
        max_zoom=20,
        max_native_zoom=20,
        detect_retina=True,
    ).add_to(m)

    # Route + buffer
    folium.FeatureGroup(name="航线", show=True).add_to(m)
    folium.PolyLine(route_line, color="#00c853", weight=4, opacity=0.95, tooltip=route_name).add_to(m)
    folium.GeoJson(
        {"type": "Feature", "geometry": mapping(buffer_wgs), "properties": {"name": "100m缓冲区"}},
        name="航线缓冲区(100m)",
        style_function=lambda _: {"color": "#ff1744", "weight": 2, "fillColor": "#ff5252", "fillOpacity": 0.22},
    ).add_to(m)

    # Population layers (one per city)
    for city, tdir in population_tile_dirs:
        tiles_uri = tdir.resolve().as_uri()
        folium.TileLayer(
            tiles=f"{tiles_uri}/{{z}}/{{x}}/{{y}}.png",
            attr=f"WorldPop Local {city}",
            name=f"人口密度-{city}",
            overlay=True,
            control=True,
            opacity=0.7,
            tms=True,
            show=False,
        ).add_to(m)

    # Landuse layer
    land_fg = folium.FeatureGroup(name="土地利用", show=False)
    if landuse_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": landuse_features},
            style_function=lambda _: {"color": "#ef6c00", "weight": 1, "fillColor": "#ffb74d", "fillOpacity": 0.2},
        ).add_to(land_fg)
    land_fg.add_to(m)

    water_surface_fg = folium.FeatureGroup(name="水面分布", show=False)
    if water_surface_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": water_surface_features},
            style_function=lambda _: {
                "color": "#0288d1",
                "weight": 1.2,
                "fillColor": "#4fc3f7",
                "fillOpacity": 0.28,
            },
        ).add_to(water_surface_fg)
    water_surface_fg.add_to(m)

    waterway_fg = folium.FeatureGroup(name="河道分布", show=False)
    if waterway_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": waterway_features},
            style_function=lambda _: {"color": "#0277bd", "weight": 2, "opacity": 0.92},
        ).add_to(waterway_fg)
    waterway_fg.add_to(m)

    # Roads layer
    roads_fg = folium.FeatureGroup(name="道路", show=False)
    if road_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": road_features},
            style_function=lambda _: {"color": "#1e88e5", "weight": 1.6, "opacity": 0.9},
        ).add_to(roads_fg)
    roads_fg.add_to(m)

    # High-speed rail layer
    hsr_fg = folium.FeatureGroup(name="高铁", show=False)
    if hsr_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": hsr_features},
            style_function=lambda _: {"color": "#e53935", "weight": 2.2, "opacity": 0.95},
        ).add_to(hsr_fg)
    hsr_fg.add_to(m)

    # POI layer
    poi_fg = folium.FeatureGroup(name="POI", show=True)
    cluster = MarkerCluster().add_to(poi_fg)
    for feat in poi_features:
        geom = feat.get("geometry", {})
        if geom.get("type") != "Point":
            continue
        coords_xy = geom.get("coordinates", [])
        if len(coords_xy) != 2:
            continue
        lon, lat = coords_xy
        props = feat.get("properties", {})
        name = props.get("name") or "未命名POI"
        ptype = props.get("poi_type") or ""
        folium.CircleMarker(
            location=[lat, lon],
            radius=3,
            color="#1565c0",
            fill=True,
            fill_color="#1e88e5",
            fill_opacity=0.9,
            weight=1,
            popup=folium.Popup(f"<b>{name}</b><br>{ptype}", max_width=280),
        ).add_to(cluster)
    poi_fg.add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)
    m.save(str(out_html))


def format_ra_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for c in col:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len * 1.2), 60)

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Workflow v2: city data + route HTML + POI/tif risk -> RA_v2")
    parser.add_argument("--kml", action="append", required=True, help="KML path/dir/glob, repeatable")
    parser.add_argument("--out-dir", default="output/full-workflow-v2", help="output root")
    parser.add_argument("--xlsx", default="RA_v2.xlsx", help="new RA workbook path")
    parser.add_argument("--city-zoom", default="8-14", help="population tile zoom for city downloader")
    parser.add_argument(
        "--poi-rules",
        default="skills/full_assessment_workflow/poi_rules.json",
        help="POI classification rules JSON path",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[2]
    out_dir = (root / args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = (root / args.xlsx).resolve()
    rules_path = (root / args.poi_rules).resolve()
    poi_rules = load_poi_rules(rules_path)

    kml_files = resolve_kmls(args.kml)
    if not kml_files:
        raise FileNotFoundError("No KML files found.")

    route_ctx: List[Dict[str, Any]] = []
    city_set = set()

    print("Step 1: Detect cities for all routes ...", flush=True)
    for kml in kml_files:
        coords = parse_kml_coords(kml)
        _, buffer_wgs = build_route_buffer(coords, BUFFER_EACH_SIDE_M)
        cities = detect_cities_for_route(coords, buffer_wgs)
        if not cities:
            # fallback: center point
            c_lat = sum(c[0] for c in coords) / len(coords)
            c_lon = sum(c[1] for c in coords) / len(coords)
            c = reverse_geocode_city(c_lat, c_lon)
            if c:
                cities = [c]
        for c in cities:
            city_set.add(c)
        route_ctx.append({
            "kml": kml,
            "coords": coords,
            "buffer": buffer_wgs,
            "cities": cities,
        })

    print(f"Detected cities: {sorted(city_set)}", flush=True)

    # Use route coverage bbox per city to avoid downloading full-city datasets for large cities.
    city_bbox_overrides: Dict[str, Tuple[float, float, float, float]] = {}
    bbox_pad = 0.02
    for rc in route_ctx:
        minx, miny, maxx, maxy = rc["buffer"].bounds
        south = miny - bbox_pad
        north = maxy + bbox_pad
        west = minx - bbox_pad
        east = maxx + bbox_pad
        for city in rc["cities"]:
            if city not in city_bbox_overrides:
                city_bbox_overrides[city] = (south, north, west, east)
                continue
            s0, n0, w0, e0 = city_bbox_overrides[city]
            city_bbox_overrides[city] = (min(s0, south), max(n0, north), min(w0, west), max(e0, east))

    print("Step 2: Ensure city datasets (skip existing) ...", flush=True)
    city_summaries: Dict[str, Dict[str, Any]] = {}
    for city in sorted(city_set):
        city_summaries[city] = ensure_city_data(
            root,
            city,
            zoom=args.city_zoom,
            bbox_override=city_bbox_overrides.get(city),
        )

    rows: List[Dict[str, Any]] = []
    summary_routes: List[Dict[str, Any]] = []

    print("Step 3~5: Build route HTML + compute risks + write RA rows ...", flush=True)

    for rc in route_ctx:
        kml: Path = rc["kml"]
        coords = rc["coords"]
        buffer_wgs: Polygon = rc["buffer"]
        route_cities: List[str] = rc["cities"]
        route_line_wgs = LineString([(c[1], c[0]) for c in coords])

        route_name = kml.stem
        route_dir = out_dir / route_name
        route_dir.mkdir(parents=True, exist_ok=True)

        # Collect city layers for this route
        poi_all_raw: List[Dict[str, Any]] = []
        landuse_all_raw: List[Dict[str, Any]] = []
        roads_all_raw: List[Dict[str, Any]] = []
        hsr_all_raw: List[Dict[str, Any]] = []
        water_surface_all_raw: List[Dict[str, Any]] = []
        waterway_all_raw: List[Dict[str, Any]] = []
        pop_tile_dirs: List[Tuple[str, Path]] = []
        pop_tifs: List[Path] = []

        for city in route_cities:
            s = city_summaries.get(city, {})
            outputs = s.get("outputs", {})
            poi_path = Path(outputs.get("poi", {}).get("geojson", ""))
            landuse_path = Path(outputs.get("landuse", {}).get("geojson", ""))
            pop_tiles = Path(outputs.get("population", {}).get("tiles_dir", ""))
            pop_tif = Path(outputs.get("population", {}).get("clipped_tif", ""))
            roads_path = Path(outputs.get("transport", {}).get("roads_geojson", ""))
            hsr_path = Path(outputs.get("transport", {}).get("hsr_geojson", ""))
            water_surface_path = Path(outputs.get("hydro", {}).get("water_surface_geojson", ""))
            waterway_path = Path(outputs.get("hydro", {}).get("waterway_geojson", ""))

            if poi_path.exists():
                poi_all_raw.extend(load_geojson_features(poi_path))
            if landuse_path.exists():
                landuse_all_raw.extend(load_geojson_features(landuse_path))
            if pop_tiles.exists():
                pop_tile_dirs.append((city, pop_tiles))
            if pop_tif.exists():
                pop_tifs.append(pop_tif)
            if roads_path.exists():
                roads_all_raw.extend(load_geojson_features(roads_path))
            if hsr_path.exists():
                hsr_all_raw.extend(load_geojson_features(hsr_path))
            if water_surface_path.exists():
                water_surface_all_raw.extend(load_geojson_features(water_surface_path))
            if waterway_path.exists():
                waterway_all_raw.extend(load_geojson_features(waterway_path))

        # Route-local filtering for performance
        minx, miny, maxx, maxy = buffer_wgs.bounds
        bbox_pad = 0.02
        route_bbox = box(minx - bbox_pad, miny - bbox_pad, maxx + bbox_pad, maxy + bbox_pad)
        prep_route_bbox = prep(route_bbox)

        poi_all: List[Dict[str, Any]] = []
        for feat in poi_all_raw:
            geom = feat.get("geometry", {})
            if geom.get("type") != "Point":
                continue
            xy = geom.get("coordinates", [])
            if len(xy) != 2:
                continue
            p = Point(float(xy[0]), float(xy[1]))
            if prep_route_bbox.contains(p) or prep_route_bbox.intersects(p):
                poi_all.append(feat)

        landuse_all: List[Dict[str, Any]] = []
        for feat in landuse_all_raw:
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if prep_route_bbox.intersects(g):
                landuse_all.append(feat)

        def _filter_line_features(features: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            out = []
            for feat in features:
                geom = feat.get("geometry")
                if not geom:
                    continue
                try:
                    g = shape(geom)
                except Exception:
                    continue
                if g.is_empty:
                    continue
                if prep_route_bbox.intersects(g):
                    out.append(feat)
            return out

        def _filter_geo_features(features: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            out = []
            for feat in features:
                geom = feat.get("geometry")
                if not geom:
                    continue
                try:
                    g = shape(geom)
                except Exception:
                    continue
                if g.is_empty:
                    continue
                if prep_route_bbox.intersects(g):
                    out.append(feat)
            return out

        roads_all = _filter_line_features(roads_all_raw)
        hsr_all = _filter_line_features(hsr_all_raw)
        water_surface_all = _filter_geo_features(water_surface_all_raw)
        waterway_all = _filter_geo_features(waterway_all_raw)

        # Route map HTML
        route_html = route_dir / f"{route_name}_map.html"
        build_route_html(
            route_html,
            route_name,
            coords,
            buffer_wgs,
            poi_all,
            landuse_all,
            water_surface_all,
            waterway_all,
            pop_tile_dirs,
            roads_all,
            hsr_all,
        )

        # POI in buffer classification
        prep_buf = prep(buffer_wgs)
        key_pts: List[str] = []
        crowd_pts: List[str] = []
        obs_pts: List[str] = []
        detail_pts: List[str] = []

        for feat in poi_all:
            geom = feat.get("geometry", {})
            if geom.get("type") != "Point":
                continue
            xy = geom.get("coordinates", [])
            if len(xy) != 2:
                continue
            lon, lat = float(xy[0]), float(xy[1])
            p = Point(lon, lat)
            if not prep_buf.contains(p) and not prep_buf.intersects(p):
                continue

            props = feat.get("properties", {})
            cat = classify_poi(props, poi_rules)
            if not cat:
                continue

            name = props.get("name") or "未命名POI"
            ptype = props.get("poi_type") or ""
            item = f"{name}({ptype})"
            detail_pts.append(f"{cat}:{item}|{lat:.6f},{lon:.6f}")
            if cat == "关键设施风险点":
                key_pts.append(item)
            elif cat == "人群聚集场所风险点":
                crowd_pts.append(item)
            elif cat == "活动障碍物风险点(塔吊)":
                obs_pts.append(item)

        # Extra rule: route crossing high-speed rail is a key facility risk point.
        hsr_cross_items = set()
        for feat in hsr_all:
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty or not route_line_wgs.intersects(g):
                continue

            props = feat.get("properties", {})
            line_name = props.get("name") or props.get("id") or "未命名高铁线"
            item = f"穿越高铁线({line_name})"
            if item in hsr_cross_items:
                continue
            hsr_cross_items.add(item)
            key_pts.append(item)

            inter = route_line_wgs.intersection(g)
            if not inter.is_empty:
                rp = inter.representative_point()
                detail_pts.append(f"关键设施风险点:{item}|{rp.y:.6f},{rp.x:.6f}")
            else:
                detail_pts.append(f"关键设施风险点:{item}|,")

        total_risk_pts = len(key_pts) + len(crowd_pts) + len(obs_pts)
        env_risk = risk_level_from_env(total_risk_pts)

        # Water polygons for population masking (prefer dedicated water-surface layer).
        water_geoms = []
        for feat in water_surface_all:
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if g.geom_type not in {"Polygon", "MultiPolygon"}:
                continue
            if not g.is_valid:
                g = g.buffer(0)
            if g.is_empty:
                continue
            water_geoms.append(g)

        for feat in landuse_all:
            if not is_water_landuse_feature(feat):
                continue
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if g.geom_type not in {"Polygon", "MultiPolygon"}:
                continue
            if not g.is_valid:
                g = g.buffer(0)
            if g.is_empty:
                continue
            water_geoms.append(g)
        water_union = unary_union(water_geoms) if water_geoms else None

        # population from tif (choose first city tif; fallback global)
        pop_tif = pop_tifs[0] if pop_tifs else (root / "data" / "population" / "chn_pd_2020_1km.tif")
        avg_pop = calc_population_with_water_zero(buffer_wgs, pop_tif, water_union)
        pop_risk = risk_level_from_population(avg_pop)

        max_alt = max(c[2] for c in coords)
        h_risk = risk_level_from_height(max_alt)
        overall = max_risk_level(h_risk, env_risk, pop_risk)

        row = {
            "记录时间": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "航线名称": kml.name,
            "航线文件链接": str(kml.resolve()),
            "地图HTML链接": str(route_html.resolve()),
            "覆盖城市": "、".join(route_cities),
            "航线最高高度(米)": round(max_alt, 2),
            "高度风险分级": h_risk,
            "关键设施风险点": "；".join(key_pts),
            "人群聚集场所风险点": "；".join(crowd_pts),
            "活动障碍物风险点(塔吊)": "；".join(obs_pts),
            "缓冲区风险点总数": total_risk_pts,
            "环境风险分级": env_risk,
            "缓冲区平均人口密度(人/平方公里)": round(avg_pop, 2),
            "人口密度风险分级": pop_risk,
            "总体风险等级": overall,
            "风险点明细": "；".join(detail_pts),
        }
        rows.append(row)

        summary_routes.append(
            {
                "route": kml.name,
                "cities": route_cities,
                "map_html": str(route_html.resolve()),
                "risk_point_total": total_risk_pts,
                "avg_pop": round(avg_pop, 2),
                "overall": overall,
            }
        )

    df = pd.DataFrame(rows).reindex(columns=RA_COLUMNS)
    df.to_excel(xlsx_path, index=False)
    format_ra_xlsx(xlsx_path)

    summary_path = out_dir / "workflow_summary_v2.json"
    summary_path.write_text(
        json.dumps(
            {
                "routes": summary_routes,
                "cities": sorted(city_set),
                "ra_xlsx": str(xlsx_path.resolve()),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Workflow v2 completed for {len(rows)} route(s)")
    print(f"RA: {xlsx_path}")
    print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
