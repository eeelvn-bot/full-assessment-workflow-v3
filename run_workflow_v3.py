#!/usr/bin/env python3
"""航线风险评估: evaluate existing routes only with hard constraints + population + weighted score."""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.util
import json
import math
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib import parse, request

import folium
import pandas as pd
import pyproj
from folium.plugins import MarkerCluster
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from osgeo import gdal
from shapely.geometry import GeometryCollection, LineString, Point, Polygon, box, mapping, shape
from shapely.ops import transform, unary_union
from shapely.prepared import prep

BUFFER_EACH_SIDE_M = 100.0
HARD_MAX_TRUE_HEIGHT_M = 200.0
RA_BREACH_DETAIL_LIMIT = 120

# V3: 3 dimensions only
DEFAULT_HARD_WEIGHT = 0.65
DEFAULT_POP_WEIGHT = 0.35
DEFAULT_HARD_NORM_CAP = 3.0
DEFAULT_POP_NORM_CAP = 20000.0

_GEOCODE_CACHE: Dict[Tuple[float, float], Optional[str]] = {}

RA_COLUMNS = [
    "记录时间",
    "航线名称",
    "航线文件链接",
    "地图HTML链接",
    "Meta链接",
    "覆盖城市",
    "航线最高高度(米)",
    "平均人口密度(人/平方公里)",
    "硬约束突破次数",
    "硬约束突破点",
    "硬约束类型统计",
    "归一化加权评分(0-100)",
    "合规结论",
    "风险等级",
    "评分明细",
    "整改状态",
    "整改建议",
    "整改后航线链接",
    "整改后Meta链接",
    "整改对比摘要",
]

REPLAN_POLICY_PRESETS: Dict[str, Dict[str, float]] = {
    "balanced": {
        "reference_corridor_m": 300.0,
        "reference_deviation_weight": 1.4,
        "reference_max_detour_ratio": 1.2,
        "reference_max_mean_offset_m": 180.0,
    },
    "strict": {
        "reference_corridor_m": 220.0,
        "reference_deviation_weight": 1.8,
        "reference_max_detour_ratio": 1.12,
        "reference_max_mean_offset_m": 130.0,
    },
    "compliance_first": {
        "reference_corridor_m": 420.0,
        "reference_deviation_weight": 0.9,
        "reference_max_detour_ratio": 1.35,
        "reference_max_mean_offset_m": 260.0,
    },
}


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def slugify(value: str) -> str:
    value = value.strip().replace(" ", "_")
    safe = []
    for ch in value:
        if ch.isalnum() or ch in "-_":
            safe.append(ch)
        else:
            safe.append("-")
    result = "".join(safe).strip("-")
    return result or "city"


def run_cmd(cmd: List[str]) -> None:
    print("RUN:", " ".join(str(c) for c in cmd), flush=True)
    subprocess.run(cmd, check=True)


def resolve_kmls(kml_args: Iterable[str]) -> List[Path]:
    files: List[Path] = []
    for item in kml_args:
        p = Path(item)
        if any(ch in item for ch in ["*", "?", "["]):
            files.extend(sorted(Path(".").glob(item)))
        elif p.is_dir():
            files.extend(sorted(p.glob("*.kml")))
        else:
            files.append(p)

    out: List[Path] = []
    seen = set()
    for f in files:
        af = f.resolve()
        if af not in seen:
            seen.add(af)
            out.append(af)
    return out


def parse_kml_coords(kml_path: Path) -> List[Tuple[float, float, float]]:
    text = kml_path.read_text(encoding="utf-8")
    m = re.search(r"<LineString>.*?<coordinates>(.*?)</coordinates>", text, re.DOTALL)
    if not m:
        m = re.search(r"<coordinates>(.*?)</coordinates>", text, re.DOTALL)
    if not m:
        raise ValueError(f"No coordinates in KML: {kml_path}")

    coords: List[Tuple[float, float, float]] = []
    for c in m.group(1).strip().split():
        parts = c.split(",")
        if len(parts) >= 2:
            lon = float(parts[0])
            lat = float(parts[1])
            alt = float(parts[2]) if len(parts) > 2 and parts[2] else 0.0
            coords.append((lat, lon, alt))
    if len(coords) < 2:
        raise ValueError(f"Insufficient coordinates in KML: {kml_path}")
    return coords


def build_projectors(coords: List[Tuple[float, float, float]]):
    center_lon = sum(c[1] for c in coords) / len(coords)
    zone = int((center_lon + 180) / 6) + 1
    south = sum(c[0] for c in coords) / len(coords) < 0
    epsg = 32700 + zone if south else 32600 + zone
    fwd = pyproj.Transformer.from_crs("EPSG:4326", f"EPSG:{epsg}", always_xy=True).transform
    inv = pyproj.Transformer.from_crs(f"EPSG:{epsg}", "EPSG:4326", always_xy=True).transform
    return fwd, inv


def build_route_buffer(coords: List[Tuple[float, float, float]], each_side_m: float) -> Tuple[LineString, Polygon]:
    line_wgs = LineString([(c[1], c[0]) for c in coords])
    fwd, inv = build_projectors(coords)
    line_proj = transform(fwd, line_wgs)
    buffer_proj = line_proj.buffer(each_side_m)
    buffer_wgs = transform(inv, buffer_proj)
    return line_wgs, buffer_wgs


def reverse_geocode_city(lat: float, lon: float) -> Optional[str]:
    key = (round(lat, 4), round(lon, 4))
    if key in _GEOCODE_CACHE:
        return _GEOCODE_CACHE[key]
    params = parse.urlencode(
        {
            "format": "jsonv2",
            "lat": lat,
            "lon": lon,
            "zoom": 8,
            "addressdetails": 1,
        }
    )
    url = f"https://nominatim.openstreetmap.org/reverse?{params}"
    req = request.Request(url, headers={"User-Agent": "risk-workflow-v3/1.0"})
    try:
        with request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        _GEOCODE_CACHE[key] = None
        return None
    addr = data.get("address", {})
    display_name = str(data.get("display_name", ""))
    if display_name:
        for part in [p.strip() for p in display_name.split(",")]:
            if part.endswith("市") and part not in {"中国"}:
                _GEOCODE_CACHE[key] = part
                return part
    city = str(addr.get("city", "")).strip()
    if city and not city.endswith(("区", "县")):
        _GEOCODE_CACHE[key] = city
        return city

    state_district = str(addr.get("state_district", "")).strip()
    if state_district and state_district.endswith("市"):
        _GEOCODE_CACHE[key] = state_district
        return state_district

    for addr_key in ["city", "county", "town", "district"]:
        val = str(addr.get(addr_key, "")).strip()
        if val:
            _GEOCODE_CACHE[key] = val
            return val
    _GEOCODE_CACHE[key] = None
    return None


def detect_cities_for_route(coords: List[Tuple[float, float, float]], buffer_wgs: Polygon) -> List[str]:
    del buffer_wgs
    sample_points: List[Tuple[float, float]] = []
    sample_points.append((coords[0][0], coords[0][1]))
    sample_points.append((coords[-1][0], coords[-1][1]))
    sample_points.append((sum(c[0] for c in coords) / len(coords), sum(c[1] for c in coords) / len(coords)))

    cities: List[str] = []
    seen = set()
    for lat, lon in sample_points:
        city = reverse_geocode_city(lat, lon)
        if city and city not in seen:
            seen.add(city)
            cities.append(city)
    return cities


def city_cache_dir(root: Path, city: str) -> Path:
    return root / "output" / "city_data_cache" / slugify(city)


def _summary_complete(summary: Dict[str, Any]) -> bool:
    outputs = summary.get("outputs", {})
    need = ["poi", "landuse", "hydro", "population", "transport"]
    return all(k in outputs for k in need)


def ensure_city_data(
    root: Path,
    city: str,
    zoom: str = "8-14",
    bbox_override: Optional[Tuple[float, float, float, float]] = None,
) -> Dict[str, Any]:
    cdir = city_cache_dir(root, city)
    cdir.mkdir(parents=True, exist_ok=True)
    summary_path = cdir / "download_summary.json"

    if summary_path.exists():
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            if _summary_complete(summary):
                print(f"[SKIP] city data exists: {city} -> {cdir}", flush=True)
                return summary
        except Exception:
            pass

    script = root / "skills" / "city-data-downloader" / "scripts" / "download_city_data.py"
    cmd = [
        "python3",
        str(script),
        "--city",
        city,
        "--outdir",
        str(cdir),
        "--zoom",
        zoom,
        "--landuse-keys",
        "landuse",
    ]
    if bbox_override is not None:
        south, north, west, east = bbox_override
        cmd.extend(["--bbox", f"{south},{north},{west},{east}"])
    run_cmd(cmd)
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    return summary


def parse_raw_tags(raw: Any) -> Dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            return json.loads(raw)
        except Exception:
            return {}
    return {}


def is_water_landuse_feature(feature: Dict[str, Any]) -> bool:
    props = feature.get("properties", {})
    ltype = str(props.get("landuse_type", "")).lower()
    tags = parse_raw_tags(props.get("raw_tags", {}))

    if ltype.startswith("natural:water") or ltype.startswith("waterway:"):
        return True
    if ltype.startswith("landuse:reservoir") or ltype.startswith("landuse:basin"):
        return True

    nat = str(tags.get("natural", "")).lower()
    waterway = str(tags.get("waterway", "")).lower()
    landuse = str(tags.get("landuse", "")).lower()
    if nat in {"water", "wetland"}:
        return True
    if waterway:
        return True
    if landuse in {"reservoir", "basin"}:
        return True
    return False


def load_geojson_features(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    obj = json.loads(path.read_text(encoding="utf-8"))
    return obj.get("features", [])


def calc_population_with_water_zero(
    buffer_wgs: Polygon,
    pop_tif: Path,
    water_union: Optional[Polygon],
) -> float:
    ds = gdal.Open(str(pop_tif))
    if ds is None:
        return 0.0
    gt = ds.GetGeoTransform()
    rb = ds.GetRasterBand(1)
    nodata = rb.GetNoDataValue()

    min_lon, min_lat, max_lon, max_lat = buffer_wgs.bounds

    px_a = (min_lon - gt[0]) / gt[1]
    px_b = (max_lon - gt[0]) / gt[1]
    py_a = (max_lat - gt[3]) / gt[5]
    py_b = (min_lat - gt[3]) / gt[5]

    px_min = math.floor(min(px_a, px_b))
    px_max = math.ceil(max(px_a, px_b))
    py_min = math.floor(min(py_a, py_b))
    py_max = math.ceil(max(py_a, py_b))

    off_x = max(0, px_min)
    off_y = max(0, py_min)
    end_x = min(ds.RasterXSize - 1, px_max)
    end_y = min(ds.RasterYSize - 1, py_max)
    win_x = end_x - off_x + 1
    win_y = end_y - off_y + 1
    if win_x <= 0 or win_y <= 0:
        return 0.0

    data = rb.ReadAsArray(off_x, off_y, win_x, win_y)
    if data is None:
        return 0.0

    prep_buf = prep(buffer_wgs)
    prep_water = prep(water_union) if water_union is not None and not water_union.is_empty else None

    weighted_sum = 0.0
    weighted_area = 0.0

    for y in range(win_y):
        for x in range(win_x):
            val = float(data[y, x])
            if nodata is not None and val == nodata:
                continue

            lon0 = gt[0] + (off_x + x) * gt[1]
            lat0 = gt[3] + (off_y + y) * gt[5]
            lon1 = lon0 + gt[1]
            lat1 = lat0 + gt[5]
            cell = box(min(lon0, lon1), min(lat0, lat1), max(lon0, lon1), max(lat0, lat1))

            if not prep_buf.intersects(cell):
                continue
            inter = cell.intersection(buffer_wgs)
            if inter.is_empty:
                continue
            inter_area = inter.area
            if inter_area <= 0:
                continue

            if prep_water is not None and prep_water.intersects(inter):
                water_area = inter.intersection(water_union).area
                land_area = max(inter_area - water_area, 0.0)
            else:
                land_area = inter_area

            weighted_sum += land_area * val
            weighted_area += inter_area

    if weighted_area <= 0:
        return 0.0
    return weighted_sum / weighted_area


def _load_plan_auto_route_module(root: Path):
    mod_path = root / "skills" / "plan-auto-route" / "scripts" / "plan_auto_route.py"
    if not mod_path.exists():
        raise FileNotFoundError(f"plan_auto_route.py not found: {mod_path}")

    mod_name = "plan_auto_route_v3_ref"
    spec = importlib.util.spec_from_file_location(mod_name, mod_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Failed to load plan_auto_route module spec")
    module = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = module
    spec.loader.exec_module(module)
    return module


def xy_geom_to_wgs_feature(geom_xy, inv, props: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if geom_xy is None:
        return None
    try:
        g_wgs = transform(inv, geom_xy)
    except Exception:
        return None
    if g_wgs.is_empty:
        return None
    return {"type": "Feature", "geometry": mapping(g_wgs), "properties": props}


def xy_geoms_to_wgs_features(geoms_xy: List[Any], inv, layer_type: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for i, g in enumerate(geoms_xy):
        feat = xy_geom_to_wgs_feature(g, inv, {"layer_type": layer_type, "id": f"{layer_type}-{i + 1}"})
        if feat is not None:
            out.append(feat)
    return out


def _collect_points_from_geom(inter_geom, inv) -> List[Tuple[float, float]]:
    out: List[Tuple[float, float]] = []
    if inter_geom is None or inter_geom.is_empty:
        return out

    def add_xy_point(x: float, y: float) -> None:
        try:
            lon, lat = inv(float(x), float(y))
        except Exception:
            return
        out.append((float(lat), float(lon)))

    def walk(g) -> None:
        gtype = g.geom_type
        if gtype == "Point":
            add_xy_point(g.x, g.y)
            return
        if gtype == "MultiPoint":
            for gg in g.geoms:
                add_xy_point(gg.x, gg.y)
            return
        if gtype in {"LineString", "LinearRing"}:
            rp = g.representative_point()
            add_xy_point(rp.x, rp.y)
            return
        if gtype == "MultiLineString":
            for gg in g.geoms:
                rp = gg.representative_point()
                add_xy_point(rp.x, rp.y)
            return
        if gtype == "Polygon":
            rp = g.representative_point()
            add_xy_point(rp.x, rp.y)
            return
        if gtype == "MultiPolygon":
            for gg in g.geoms:
                rp = gg.representative_point()
                add_xy_point(rp.x, rp.y)
            return
        if isinstance(g, GeometryCollection) or gtype == "GeometryCollection":
            for gg in g.geoms:
                walk(gg)

    walk(inter_geom)

    dedup = {}
    for lat, lon in out:
        dedup[(round(lat, 6), round(lon, 6))] = (lat, lon)
    return list(dedup.values())


def _pick_event_anchor(inter_geom, fallback_geom, inv) -> Tuple[Optional[float], Optional[float]]:
    pts = _collect_points_from_geom(inter_geom, inv)
    if not pts:
        pts = _collect_points_from_geom(fallback_geom, inv)
    if not pts:
        return None, None
    return pts[0][0], pts[0][1]


def _add_hard_event(
    events: Dict[str, Dict[str, Any]],
    event_key: str,
    ctype: str,
    detail: str,
    lat: Optional[float],
    lon: Optional[float],
) -> None:
    if event_key in events:
        return
    events[event_key] = {
        "event_key": event_key,
        "type": ctype,
        "detail": detail,
        "lat": lat,
        "lon": lon,
    }


def evaluate_hard_constraints(
    coords: List[Tuple[float, float, float]],
    route_line_xy,
    route_buffer_xy,
    inv,
    civil_hard_polys_xy: List[Any],
    military_hard_polys_xy: List[Any],
    school_hard_union_xy,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    events: Dict[str, Dict[str, Any]] = {}

    # Civil no-fly: use polygon boundary overlap against route buffer (requested strict rule).
    for i, poly in enumerate(civil_hard_polys_xy):
        try:
            if poly.is_empty or (not route_buffer_xy.intersects(poly)):
                continue
            inter = route_buffer_xy.intersection(poly)
            lat, lon = _pick_event_anchor(inter, poly.representative_point(), inv)
            _add_hard_event(
                events=events,
                event_key=f"civil_nofly:{i + 1}",
                ctype="民航禁飞区(硬约束)",
                detail=f"民航机场禁飞区#{i + 1}与100m缓冲区重合",
                lat=lat,
                lon=lon,
            )
        except Exception:
            continue

    # Military airport hard no-fly: intersection against route centerline.
    for i, poly in enumerate(military_hard_polys_xy):
        try:
            if poly.is_empty or (not route_line_xy.intersects(poly)):
                continue
            inter = route_line_xy.intersection(poly)
            lat, lon = _pick_event_anchor(inter, poly.representative_point(), inv)
            _add_hard_event(
                events=events,
                event_key=f"military_nofly:{i + 1}",
                ctype="军用机场禁飞区(硬约束)",
                detail=f"军用机场禁飞区#{i + 1}与航线重合",
                lat=lat,
                lon=lon,
            )
        except Exception:
            continue

    # School/kindergarten hard constraint from efficiency route standard.
    if school_hard_union_xy is not None and (not school_hard_union_xy.is_empty):
        try:
            if route_line_xy.intersects(school_hard_union_xy):
                inter = route_line_xy.intersection(school_hard_union_xy)
                lat, lon = _pick_event_anchor(inter, school_hard_union_xy.representative_point(), inv)
                _add_hard_event(
                    events=events,
                    event_key="school_kindergarten_hard:route_overlap",
                    ctype="学校/幼儿园避让区(硬约束)",
                    detail="学校/幼儿园避让区与航线重合",
                    lat=lat,
                    lon=lon,
                )
        except Exception:
            pass

    # Max true-height hard cap: de-duplicate as one event with waypoint count.
    exceed_points: List[Tuple[int, float, float, float]] = []
    for i, (lat, lon, alt) in enumerate(coords):
        alt_f = float(alt)
        if alt_f <= HARD_MAX_TRUE_HEIGHT_M:
            continue
        exceed_points.append((i + 1, float(lat), float(lon), alt_f))
    if exceed_points:
        max_pt = max(exceed_points, key=lambda x: x[3])
        _add_hard_event(
            events=events,
            event_key="max_true_height:exceed_200m",
            ctype="最大真高超限(硬约束)",
            detail=(
                f"超限航点{len(exceed_points)}个，"
                f"最高航点#{max_pt[0]} 高度{max_pt[3]:.2f}m > {HARD_MAX_TRUE_HEIGHT_M:.0f}m"
            ),
            lat=max_pt[1],
            lon=max_pt[2],
        )

    records = sorted(events.values(), key=lambda x: (str(x.get("type", "")), str(x.get("detail", ""))))
    type_counter: Dict[str, int] = {}
    for r in records:
        ctype = str(r.get("type", ""))
        type_counter[ctype] = int(type_counter.get(ctype, 0) + 1)

    return records, type_counter


def compute_weighted_score(
    hard_break_count: int,
    avg_pop: float,
    w_hard: float,
    w_pop: float,
    hard_norm_cap: float,
    pop_norm_cap: float,
) -> Tuple[float, Dict[str, float]]:
    hard_norm = clamp(float(hard_break_count) / max(1e-6, hard_norm_cap), 0.0, 1.0)
    pop_norm = clamp(float(avg_pop) / max(1e-6, pop_norm_cap), 0.0, 1.0)
    weighted = 100.0 * (w_hard * hard_norm + w_pop * pop_norm)
    return round(weighted, 2), {
        "hard_norm": round(hard_norm, 6),
        "pop_norm": round(pop_norm, 6),
        "w_hard": round(w_hard, 6),
        "w_pop": round(w_pop, 6),
    }


def serialize_breach_details(records: List[Dict[str, Any]], limit: int = RA_BREACH_DETAIL_LIMIT) -> str:
    if not records:
        return ""
    parts: List[str] = []
    max_n = min(limit, len(records))
    for r in records[:max_n]:
        lat = r.get("lat")
        lon = r.get("lon")
        if lat is None or lon is None:
            coord = ","
        else:
            coord = f"{float(lat):.6f},{float(lon):.6f}"
        parts.append(f"{r.get('type', '')}:{r.get('detail', '')}|{coord}")
    if len(records) > limit:
        parts.append(f"...已截断，总计{len(records)}处")
    return "；".join(parts)


def severity_from_metrics(hard_break_count: int, avg_pop: float) -> str:
    if hard_break_count > 0:
        return "高"
    if avg_pop > 20000:
        return "高"
    if avg_pop > 5000:
        return "中"
    return "低"


def _sample_ratios_for_line(length_m: float) -> List[float]:
    if length_m <= 120.0:
        return [0.2, 0.5, 0.8]
    if length_m <= 360.0:
        return [0.12, 0.3, 0.5, 0.7, 0.88]
    if length_m <= 1200.0:
        return [0.08, 0.2, 0.32, 0.44, 0.56, 0.68, 0.8, 0.92]
    return [0.05, 0.14, 0.23, 0.32, 0.41, 0.5, 0.59, 0.68, 0.77, 0.86, 0.95]


def route_length_m(coords: List[Tuple[float, float, float]]) -> float:
    if len(coords) < 2:
        return 0.0
    line_wgs = LineString([(c[1], c[0]) for c in coords])
    fwd, _inv = build_projectors(coords)
    return float(transform(fwd, line_wgs).length)


def route_offset_stats_m(
    reference_coords: List[Tuple[float, float, float]],
    new_coords: List[Tuple[float, float, float]],
) -> Tuple[float, float, float]:
    if len(reference_coords) < 2 or len(new_coords) < 2:
        return 0.0, 0.0, 0.0
    line_ref_wgs = LineString([(c[1], c[0]) for c in reference_coords])
    line_new_wgs = LineString([(c[1], c[0]) for c in new_coords])
    fwd, _inv = build_projectors(reference_coords + new_coords)
    line_ref_xy = transform(fwd, line_ref_wgs)
    line_new_xy = transform(fwd, line_new_wgs)
    if line_new_xy.is_empty or line_ref_xy.is_empty:
        return 0.0, 0.0, 0.0
    vals: List[float] = []
    for ratio in _sample_ratios_for_line(float(line_new_xy.length)):
        p = line_new_xy.interpolate(float(line_new_xy.length) * ratio)
        vals.append(float(p.distance(line_ref_xy)))
    if not vals:
        return 0.0, 0.0, 0.0
    vals_sorted = sorted(vals)
    q = 0.9 * (len(vals_sorted) - 1)
    lo = int(math.floor(q))
    hi = int(math.ceil(q))
    if lo == hi:
        p90 = float(vals_sorted[lo])
    else:
        w = q - lo
        p90 = float(vals_sorted[lo] * (1.0 - w) + vals_sorted[hi] * w)
    return float(sum(vals) / len(vals)), p90, float(max(vals))


def choose_primary_city(route_cities: List[str], coords: List[Tuple[float, float, float]]) -> Optional[str]:
    if not route_cities:
        return None
    if len(route_cities) == 1:
        return route_cities[0]
    mid = coords[len(coords) // 2]
    mid_city = reverse_geocode_city(float(mid[0]), float(mid[1]))
    if mid_city and mid_city in route_cities:
        return mid_city
    return route_cities[0]


def build_route_html(
    out_html: Path,
    route_name: str,
    coords: List[Tuple[float, float, float]],
    buffer_wgs: Polygon,
    poi_features: List[Dict[str, Any]],
    landuse_features: List[Dict[str, Any]],
    water_surface_features: List[Dict[str, Any]],
    waterway_features: List[Dict[str, Any]],
    population_tile_dirs: List[Tuple[str, Path]],
    road_features: List[Dict[str, Any]],
    hsr_features: List[Dict[str, Any]],
    civil_nofly_features: List[Dict[str, Any]],
    military_nofly_features: List[Dict[str, Any]],
    school_hard_features: List[Dict[str, Any]],
    hv_line_features: List[Dict[str, Any]],
    hard_breach_points: List[Dict[str, Any]],
    plan_mod: Any,
    replan_coords: Optional[List[Tuple[float, float, float]]] = None,
    replan_buffer_wgs: Optional[Polygon] = None,
    replan_hard_breach_points: Optional[List[Dict[str, Any]]] = None,
    replan_summary: Optional[Dict[str, Any]] = None,
) -> None:
    center_lat = sum(c[0] for c in coords) / len(coords)
    center_lon = sum(c[1] for c in coords) / len(coords)
    route_line = [(c[0], c[1]) for c in coords]
    route_alt_min = min(c[2] for c in coords)
    route_alt_max = max(c[2] for c in coords)
    route_variant_points = [{"lon": float(c[1]), "lat": float(c[0]), "alt": float(c[2])} for c in coords]

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=13,
        tiles=None,
        control_scale=True,
        prefer_canvas=True,
        max_zoom=22,
    )

    folium.TileLayer(
        tiles="OpenStreetMap",
        name="普通地图",
        overlay=False,
        control=True,
        show=True,
        max_native_zoom=19,
        max_zoom=22,
    ).add_to(m)
    folium.TileLayer(
        tiles="https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
        attr='&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors '
        '&copy; <a href="https://carto.com/">CARTO</a>',
        name="浅色底图",
        overlay=False,
        control=True,
        show=False,
        max_native_zoom=20,
        max_zoom=20,
    ).add_to(m)
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Tiles &copy; Esri",
        name="卫星影像",
        overlay=False,
        control=True,
        show=False,
        max_native_zoom=19,
        max_zoom=22,
    ).add_to(m)
    folium.TileLayer(
        tiles="https://services.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}",
        attr="Labels &copy; Esri",
        name="卫星注记",
        overlay=True,
        control=True,
        show=False,
        opacity=0.9,
        max_native_zoom=19,
        max_zoom=22,
    ).add_to(m)

    route_fg = folium.FeatureGroup(name="安全优先（3D高度）", show=True)
    folium.PolyLine(
        route_line,
        color="#1565c0",
        weight=4,
        opacity=0.92,
        tooltip=f"{route_name} | 高度最小/最大 {route_alt_min:.1f}/{route_alt_max:.1f} m",
    ).add_to(route_fg)
    route_step = max(1, len(coords) // 120)
    for i in range(0, len(coords), route_step):
        lat, lon, alt = coords[i]
        folium.CircleMarker(
            [float(lat), float(lon)],
            radius=2,
            color="#0d47a1",
            fill=True,
            fill_opacity=0.8,
            tooltip=f"高度 {float(alt):.1f} m",
        ).add_to(route_fg)
    route_fg.add_to(m)

    route_buffer_fg = folium.FeatureGroup(name=f"安全优先 缓冲区 {int(BUFFER_EACH_SIDE_M)}m", show=True)
    folium.GeoJson(
        {"type": "Feature", "geometry": mapping(buffer_wgs), "properties": {"name": "100m缓冲区"}},
        style_function=lambda _: {"color": "#1e88e5", "weight": 2, "fillColor": "#1e88e5", "fillOpacity": 0.08},
    ).add_to(route_buffer_fg)
    route_buffer_fg.add_to(m)

    replan_variant_points: List[Dict[str, float]] = []
    if replan_coords and len(replan_coords) >= 2:
        replan_route_line = [(c[0], c[1]) for c in replan_coords]
        replan_alt_min = min(c[2] for c in replan_coords)
        replan_alt_max = max(c[2] for c in replan_coords)
        replan_variant_points = [{"lon": float(c[1]), "lat": float(c[0]), "alt": float(c[2])} for c in replan_coords]
        summary_parts: List[str] = []
        if replan_summary:
            comp = replan_summary.get("comparison", {}) if isinstance(replan_summary.get("comparison", {}), dict) else {}
            new_eval = (
                replan_summary.get("new_evaluation", {})
                if isinstance(replan_summary.get("new_evaluation", {}), dict)
                else {}
            )
            if comp:
                summary_parts.append(f"绕行比(相对原航线): {comp.get('detour_ratio_vs_old', '-')}")
                summary_parts.append(f"平均偏移: {comp.get('mean_offset_m', '-')}m")
            if new_eval:
                summary_parts.append(f"整改后结论: {new_eval.get('compliance', '-')}")
        replan_tip = (
            f"{route_name}整改后 | 高度最小/最大 {replan_alt_min:.1f}/{replan_alt_max:.1f} m"
            + (f" | {' | '.join(summary_parts)}" if summary_parts else "")
        )

        replan_fg = folium.FeatureGroup(name="整改航线（最小改动）", show=True)
        folium.PolyLine(
            replan_route_line,
            color="#2e7d32",
            weight=4,
            opacity=0.92,
            tooltip=replan_tip,
        ).add_to(replan_fg)
        replan_step = max(1, len(replan_coords) // 120)
        for i in range(0, len(replan_coords), replan_step):
            lat, lon, alt = replan_coords[i]
            folium.CircleMarker(
                [float(lat), float(lon)],
                radius=2,
                color="#1b5e20",
                fill=True,
                fill_opacity=0.82,
                tooltip=f"整改后高度 {float(alt):.1f} m",
            ).add_to(replan_fg)
        replan_fg.add_to(m)

        if replan_buffer_wgs is not None:
            replan_buffer_fg = folium.FeatureGroup(name=f"整改航线 缓冲区 {int(BUFFER_EACH_SIDE_M)}m", show=False)
            folium.GeoJson(
                {"type": "Feature", "geometry": mapping(replan_buffer_wgs), "properties": {"name": "整改后100m缓冲区"}},
                style_function=lambda _: {"color": "#43a047", "weight": 2, "fillColor": "#66bb6a", "fillOpacity": 0.08},
            ).add_to(replan_buffer_fg)
            replan_buffer_fg.add_to(m)

    for city, tdir in population_tile_dirs:
        tiles_uri = tdir.resolve().as_uri()
        folium.TileLayer(
            tiles=f"{tiles_uri}/{{z}}/{{x}}/{{y}}.png",
            attr=f"WorldPop Local {city}",
            name=f"人口密度-{city}",
            overlay=True,
            control=True,
            opacity=0.7,
            tms=True,
            show=False,
            max_zoom=22,
        ).add_to(m)

    land_fg = folium.FeatureGroup(name="土地利用", show=False)
    if landuse_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": landuse_features},
            style_function=lambda _: {"color": "#ef6c00", "weight": 1, "fillColor": "#ffb74d", "fillOpacity": 0.2},
        ).add_to(land_fg)
    land_fg.add_to(m)

    water_surface_fg = folium.FeatureGroup(name="水面分布", show=False)
    if water_surface_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": water_surface_features},
            style_function=lambda _: {
                "color": "#0288d1",
                "weight": 1.2,
                "fillColor": "#4fc3f7",
                "fillOpacity": 0.28,
            },
        ).add_to(water_surface_fg)
    water_surface_fg.add_to(m)

    waterway_fg = folium.FeatureGroup(name="河道分布", show=False)
    if waterway_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": waterway_features},
            style_function=lambda _: {"color": "#0277bd", "weight": 2, "opacity": 0.92},
        ).add_to(waterway_fg)
    waterway_fg.add_to(m)

    line_fg = folium.FeatureGroup(name="线性风险（高速/高铁/高压电力线）", show=False)
    if road_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": road_features},
            style_function=lambda _: {"color": "#1e88e5", "weight": 1.6, "opacity": 0.9},
        ).add_to(line_fg)

    if hsr_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": hsr_features},
            style_function=lambda _: {"color": "#e53935", "weight": 2.2, "opacity": 0.95},
        ).add_to(line_fg)

    if hv_line_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": hv_line_features},
            style_function=lambda _: {"color": "#8d6e63", "weight": 2.0, "opacity": 0.95},
        ).add_to(line_fg)
    line_fg.add_to(m)

    civil_fg = folium.FeatureGroup(name="民航机场禁飞区（CAAC）", show=True)
    if civil_nofly_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": civil_nofly_features},
            style_function=lambda _: {"color": "#ad1457", "weight": 2.0, "fillColor": "#ec407a", "fillOpacity": 0.2},
        ).add_to(civil_fg)
    civil_fg.add_to(m)

    military_fg = folium.FeatureGroup(name="军用机场禁飞区（硬约束）", show=True)
    if military_nofly_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": military_nofly_features},
            style_function=lambda _: {"color": "#c62828", "weight": 2.0, "fillColor": "#ef5350", "fillOpacity": 0.2},
        ).add_to(military_fg)
    military_fg.add_to(m)

    school_fg = folium.FeatureGroup(name="学校/幼儿园避让区", show=False)
    if school_hard_features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": school_hard_features},
            style_function=lambda _: {"color": "#6a1b9a", "weight": 1.6, "fillColor": "#ab47bc", "fillOpacity": 0.18},
        ).add_to(school_fg)
    school_fg.add_to(m)

    poi_fg = folium.FeatureGroup(name="POI", show=False)
    cluster = MarkerCluster().add_to(poi_fg)
    for feat in poi_features:
        geom = feat.get("geometry", {})
        if geom.get("type") != "Point":
            continue
        coords_xy = geom.get("coordinates", [])
        if len(coords_xy) != 2:
            continue
        lon, lat = coords_xy
        props = feat.get("properties", {})
        name = props.get("name") or "未命名POI"
        ptype = props.get("poi_type") or ""
        folium.CircleMarker(
            location=[lat, lon],
            radius=3,
            color="#1565c0",
            fill=True,
            fill_color="#1e88e5",
            fill_opacity=0.9,
            weight=1,
            popup=folium.Popup(f"<b>{name}</b><br>{ptype}", max_width=280),
        ).add_to(cluster)
    poi_fg.add_to(m)

    breach_fg = folium.FeatureGroup(name="硬约束突破点", show=True)
    for p in hard_breach_points:
        lat = p.get("lat")
        lon = p.get("lon")
        if lat is None or lon is None:
            continue
        title = str(p.get("type", "硬约束突破"))
        detail = str(p.get("detail", ""))
        folium.CircleMarker(
            location=[float(lat), float(lon)],
            radius=5,
            color="#b71c1c",
            fill=True,
            fill_color="#d32f2f",
            fill_opacity=0.95,
            weight=2,
            popup=folium.Popup(f"<b>{title}</b><br>{detail}", max_width=360),
        ).add_to(breach_fg)
    breach_fg.add_to(m)

    if replan_hard_breach_points:
        replan_breach_fg = folium.FeatureGroup(name="整改后硬约束突破点", show=False)
        for p in replan_hard_breach_points:
            lat = p.get("lat")
            lon = p.get("lon")
            if lat is None or lon is None:
                continue
            title = str(p.get("type", "整改后硬约束突破"))
            detail = str(p.get("detail", ""))
            folium.CircleMarker(
                location=[float(lat), float(lon)],
                radius=5,
                color="#e65100",
                fill=True,
                fill_color="#ff9800",
                fill_opacity=0.95,
                weight=2,
                popup=folium.Popup(f"<b>{title}</b><br>{detail}", max_width=360),
            ).add_to(replan_breach_fg)
        replan_breach_fg.add_to(m)

    start_end_fg = folium.FeatureGroup(name="起点/终点", show=True)
    start_lat, start_lon, _ = coords[0]
    end_lat, end_lon, _ = coords[-1]
    folium.Marker([float(start_lat), float(start_lon)], tooltip="起点", icon=folium.Icon(color="green")).add_to(start_end_fg)
    folium.Marker([float(end_lat), float(end_lon)], tooltip="终点", icon=folium.Icon(color="red")).add_to(start_end_fg)
    start_end_fg.add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)
    html_text = m.get_root().render()
    theme_builder = getattr(plan_mod, "_build_preview_theme_head_html", None)
    if callable(theme_builder) and "</head>" in html_text:
        html_text = html_text.replace("</head>", theme_builder() + "\n</head>")
    toolbar_builder = getattr(plan_mod, "_build_preview_toolbar_script_html", None)
    if callable(toolbar_builder) and "</html>" in html_text:
        route_variants = {
            "safety_default": {
                "label": "安全优先（评估基线）",
                "points": route_variant_points,
            }
        }
        if replan_variant_points:
            route_variants["replan_min_change"] = {
                "label": "整改后（最小改动）",
                "points": replan_variant_points,
            }
        html_text = html_text.replace(
            "</html>",
            toolbar_builder(
                name=route_name,
                route_variants=route_variants,
                default_route_variant_id="safety_default",
            )
            + "\n</html>",
        )
    out_html.write_text(html_text, encoding="utf-8")


def format_ra_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for c in col:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len * 1.2), 64)

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="航线风险评估（仅评估）: hard-constraints + population + normalized weighted score"
    )
    parser.add_argument("--kml", action="append", required=True, help="KML path/dir/glob, repeatable")
    parser.add_argument("--out-dir", default="output/full-workflow-v3", help="output root")
    parser.add_argument("--xlsx", default="RA_v3.xlsx", help="new RA workbook path")
    parser.add_argument("--city-zoom", default="8-14", help="population tile zoom for city downloader")
    parser.add_argument(
        "--civil-airport-no-fly-geojson",
        default="skills/plan-auto-route/config/civil_airport_no_fly.geojson",
        help="CAAC civil airport no-fly dataset (polygon source)",
    )
    parser.add_argument("--weight-hard", type=float, default=DEFAULT_HARD_WEIGHT)
    parser.add_argument("--weight-pop", type=float, default=DEFAULT_POP_WEIGHT)
    parser.add_argument("--hard-norm-cap", type=float, default=DEFAULT_HARD_NORM_CAP)
    parser.add_argument("--pop-norm-cap", type=float, default=DEFAULT_POP_NORM_CAP)
    parser.add_argument("--route-bbox-margin-m", type=float, default=4500.0)
    parser.add_argument("--route-bbox-building-margin-m", type=float, default=2800.0)
    parser.add_argument(
        "--replan-on-noncompliant",
        choices=["off", "confirm", "auto"],
        default="off",
        help="For non-compliant route: off/confirm/auto minimal-change replanning.",
    )
    parser.add_argument(
        "--replan-policy",
        choices=["balanced", "strict", "compliance_first"],
        default="balanced",
        help="Minimal-change policy preset for reference-route replanning.",
    )
    parser.add_argument(
        "--replan-output-dir",
        default="",
        help="Optional output root for replanned route files; default is <route_dir>/replan.",
    )
    parser.add_argument(
        "--replan-select-candidate",
        choices=["safety_default", "efficiency"],
        default="safety_default",
        help="Candidate preference when replanning.",
    )
    parser.add_argument(
        "--replan-max-detour-ratio",
        type=float,
        default=None,
        help="Optional override: max detour ratio vs reference route.",
    )
    parser.add_argument(
        "--replan-max-mean-offset-m",
        type=float,
        default=None,
        help="Optional override: max mean offset (m) vs reference route.",
    )
    parser.add_argument(
        "--replan-reference-corridor-m",
        type=float,
        default=None,
        help="Optional override: normalization corridor (m) for reference-route offset.",
    )
    parser.add_argument(
        "--replan-reference-deviation-weight",
        type=float,
        default=None,
        help="Optional override: reference-route deviation weight.",
    )
    parser.add_argument(
        "--replan-profile",
        choices=["fastest", "balanced", "safest"],
        default="safest",
        help="Profile passed to plan-auto-route during replanning.",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[2]
    out_dir = (root / args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = (root / args.xlsx).resolve()
    civil_airport_geojson = (root / args.civil_airport_no_fly_geojson).resolve()

    w_hard = float(args.weight_hard)
    w_pop = float(args.weight_pop)
    w_sum = w_hard + w_pop
    if w_sum <= 0:
        raise ValueError("weight-hard + weight-pop must be > 0")
    w_hard /= w_sum
    w_pop /= w_sum

    plan_mod = _load_plan_auto_route_module(root)

    kml_files = resolve_kmls(args.kml)
    if not kml_files:
        raise FileNotFoundError("No KML files found.")
    if args.replan_on_noncompliant != "off" and len(kml_files) != 1:
        raise ValueError("整改模式仅支持单条航线输入：请只提供一个 KML。")

    route_ctx: List[Dict[str, Any]] = []
    city_set = set()

    print("Step 1: Detect cities for all routes ...", flush=True)
    for kml in kml_files:
        coords = parse_kml_coords(kml)
        _, buffer_wgs = build_route_buffer(coords, BUFFER_EACH_SIDE_M)
        cities = detect_cities_for_route(coords, buffer_wgs)
        if not cities:
            c_lat = sum(c[0] for c in coords) / len(coords)
            c_lon = sum(c[1] for c in coords) / len(coords)
            c = reverse_geocode_city(c_lat, c_lon)
            if c:
                cities = [c]
        for c in cities:
            city_set.add(c)
        route_ctx.append(
            {
                "kml": kml,
                "coords": coords,
                "buffer": buffer_wgs,
                "cities": cities,
            }
        )

    print(f"Detected cities: {sorted(city_set)}", flush=True)

    city_bbox_overrides: Dict[str, Tuple[float, float, float, float]] = {}
    bbox_pad = 0.02
    for rc in route_ctx:
        minx, miny, maxx, maxy = rc["buffer"].bounds
        south = miny - bbox_pad
        north = maxy + bbox_pad
        west = minx - bbox_pad
        east = maxx + bbox_pad
        for city in rc["cities"]:
            if city not in city_bbox_overrides:
                city_bbox_overrides[city] = (south, north, west, east)
                continue
            s0, n0, w0, e0 = city_bbox_overrides[city]
            city_bbox_overrides[city] = (min(s0, south), max(n0, north), min(w0, west), max(e0, east))

    print("Step 2: Ensure city datasets (skip existing) ...", flush=True)
    city_summaries: Dict[str, Dict[str, Any]] = {}
    for city in sorted(city_set):
        city_summaries[city] = ensure_city_data(
            root,
            city,
            zoom=args.city_zoom,
            bbox_override=city_bbox_overrides.get(city),
        )

    rows: List[Dict[str, Any]] = []
    summary_routes: List[Dict[str, Any]] = []

    print("Step 3: Evaluate each existing route (no planning) ...", flush=True)

    for rc in route_ctx:
        kml: Path = rc["kml"]
        coords = rc["coords"]
        buffer_wgs: Polygon = rc["buffer"]
        route_cities: List[str] = rc["cities"]
        route_line_wgs = LineString([(c[1], c[0]) for c in coords])

        route_name = kml.stem
        route_dir = out_dir / route_name
        route_dir.mkdir(parents=True, exist_ok=True)
        replan_info: Dict[str, Any] = {
            "attempted": False,
            "confirmed": False,
            "status": "not_triggered",
            "reason": "",
            "policy": {},
            "primary_city": "",
            "cross_city_warning": "",
            "outputs": {},
            "comparison": {},
            "new_evaluation": {},
        }
        replan_coords_for_map: Optional[List[Tuple[float, float, float]]] = None
        replan_buffer_for_map: Optional[Polygon] = None
        replan_hard_break_points_for_map: List[Dict[str, Any]] = []

        poi_all_raw: List[Dict[str, Any]] = []
        landuse_all_raw: List[Dict[str, Any]] = []
        roads_all_raw: List[Dict[str, Any]] = []
        hsr_all_raw: List[Dict[str, Any]] = []
        water_surface_all_raw: List[Dict[str, Any]] = []
        waterway_all_raw: List[Dict[str, Any]] = []
        pop_tile_dirs: List[Tuple[str, Path]] = []
        pop_tifs: List[Path] = []

        for city in route_cities:
            s = city_summaries.get(city, {})
            outputs = s.get("outputs", {})
            poi_path = Path(outputs.get("poi", {}).get("geojson", ""))
            landuse_path = Path(outputs.get("landuse", {}).get("geojson", ""))
            pop_tiles = Path(outputs.get("population", {}).get("tiles_dir", ""))
            pop_tif = Path(outputs.get("population", {}).get("clipped_tif", ""))
            roads_path = Path(outputs.get("transport", {}).get("roads_geojson", ""))
            hsr_path = Path(outputs.get("transport", {}).get("hsr_geojson", ""))
            water_surface_path = Path(outputs.get("hydro", {}).get("water_surface_geojson", ""))
            waterway_path = Path(outputs.get("hydro", {}).get("waterway_geojson", ""))

            if poi_path.exists():
                poi_all_raw.extend(load_geojson_features(poi_path))
            if landuse_path.exists():
                landuse_all_raw.extend(load_geojson_features(landuse_path))
            if pop_tiles.exists():
                pop_tile_dirs.append((city, pop_tiles))
            if pop_tif.exists():
                pop_tifs.append(pop_tif)
            if roads_path.exists():
                roads_all_raw.extend(load_geojson_features(roads_path))
            if hsr_path.exists():
                hsr_all_raw.extend(load_geojson_features(hsr_path))
            if water_surface_path.exists():
                water_surface_all_raw.extend(load_geojson_features(water_surface_path))
            if waterway_path.exists():
                waterway_all_raw.extend(load_geojson_features(waterway_path))

        minx, miny, maxx, maxy = buffer_wgs.bounds
        local_pad = 0.02
        route_bbox_poly = box(minx - local_pad, miny - local_pad, maxx + local_pad, maxy + local_pad)
        prep_route_bbox = prep(route_bbox_poly)

        poi_all: List[Dict[str, Any]] = []
        for feat in poi_all_raw:
            geom = feat.get("geometry", {})
            if geom.get("type") != "Point":
                continue
            xy = geom.get("coordinates", [])
            if len(xy) != 2:
                continue
            p = Point(float(xy[0]), float(xy[1]))
            if prep_route_bbox.contains(p) or prep_route_bbox.intersects(p):
                poi_all.append(feat)

        landuse_all: List[Dict[str, Any]] = []
        for feat in landuse_all_raw:
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if prep_route_bbox.intersects(g):
                landuse_all.append(feat)

        def _filter_geo_features(features: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            out = []
            for feat in features:
                geom = feat.get("geometry")
                if not geom:
                    continue
                try:
                    g = shape(geom)
                except Exception:
                    continue
                if g.is_empty:
                    continue
                if prep_route_bbox.intersects(g):
                    out.append(feat)
            return out

        roads_all = _filter_geo_features(roads_all_raw)
        hsr_all = _filter_geo_features(hsr_all_raw)
        water_surface_all = _filter_geo_features(water_surface_all_raw)
        waterway_all = _filter_geo_features(waterway_all_raw)

        # Water polygons for population masking.
        water_geoms = []
        for feat in water_surface_all:
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if g.geom_type not in {"Polygon", "MultiPolygon"}:
                continue
            if not g.is_valid:
                g = g.buffer(0)
            if g.is_empty:
                continue
            water_geoms.append(g)

        for feat in landuse_all:
            if not is_water_landuse_feature(feat):
                continue
            geom = feat.get("geometry")
            if not geom:
                continue
            try:
                g = shape(geom)
            except Exception:
                continue
            if g.is_empty:
                continue
            if g.geom_type not in {"Polygon", "MultiPolygon"}:
                continue
            if not g.is_valid:
                g = g.buffer(0)
            if g.is_empty:
                continue
            water_geoms.append(g)
        water_union = unary_union(water_geoms) if water_geoms else None

        pop_tif = pop_tifs[0] if pop_tifs else (root / "data" / "population" / "chn_pd_2020_1km.tif")
        avg_pop = calc_population_with_water_zero(buffer_wgs, pop_tif, water_union)
        max_alt = max(c[2] for c in coords)

        points_wgs_lonlat = [(c[1], c[0]) for c in coords]
        fwd, inv = plan_mod.build_projectors(points_wgs_lonlat)
        route_line_xy = transform(fwd, route_line_wgs)
        route_buffer_xy = transform(fwd, buffer_wgs)
        start_xy = plan_mod.wgs_to_xy(fwd, points_wgs_lonlat[0][0], points_wgs_lonlat[0][1])
        end_xy = plan_mod.wgs_to_xy(fwd, points_wgs_lonlat[-1][0], points_wgs_lonlat[-1][1])
        route_bbox_wgs = plan_mod.route_bbox_wgs(points_wgs_lonlat, margin_m=float(args.route_bbox_margin_m))
        route_bbox_building_wgs = plan_mod.route_bbox_wgs(points_wgs_lonlat, margin_m=float(args.route_bbox_building_margin_m))

        # Open-source constraints/risk data from plan-auto-route sources.
        (
            _nofly_hard_all_xy,
            nofly_soft_polys_xy,
            nofly_counter,
            civil_hard_polys_xy,
            military_hard_polys_xy,
            heli_soft_polys_xy,
        ) = plan_mod.fetch_open_data_no_fly_zones(
            route_bbox_wgs,
            fwd,
            civil_airport_geojson=str(civil_airport_geojson),
        )
        school_hard_zones_xy, school_points_xy, school_counter, _school_poi_items = plan_mod.fetch_school_kindergarten_zones(
            route_bbox_wgs,
            fwd,
        )

        school_hard_union_xy = unary_union(school_hard_zones_xy) if school_hard_zones_xy else None
        if school_hard_union_xy is not None and (not school_hard_union_xy.is_empty):
            relief_m = float(getattr(plan_mod, "SCHOOL_ENDPOINT_RELIEF_M", 55.0))
            relief = Point(start_xy).buffer(relief_m).union(Point(end_xy).buffer(relief_m))
            try:
                school_hard_union_xy = school_hard_union_xy.difference(relief)
            except Exception:
                pass
            if school_hard_union_xy.is_empty:
                school_hard_union_xy = None

        # Extra open-source layers for completeness (same sources as planner).
        road_lines_xy: List[Any] = []
        hsr_lines_xy: List[Any] = []
        hv_power_lines_xy: List[Any] = []
        line_risk_counter_total = {"highway": 0, "hsr": 0, "high_voltage_power_line": 0}
        for city in route_cities:
            summary = city_summaries.get(city, {})
            roads_xy_city, _line_union_city, hsr_xy_city, hv_xy_city, line_counter_city = plan_mod.build_line_risk_geometries(
                summary,
                route_bbox_wgs,
                fwd,
            )
            road_lines_xy.extend(roads_xy_city)
            hsr_lines_xy.extend(hsr_xy_city)
            hv_power_lines_xy.extend(hv_xy_city)
            for k, v in line_counter_city.items():
                line_risk_counter_total[k] = int(line_risk_counter_total.get(k, 0) + int(v))

        _b_geoms, _b_heights, _o_geoms, _o_heights, obstacle_counter = plan_mod.fetch_osm_buildings_and_obstacles(
            route_bbox_building_wgs,
            fwd,
        )

        hard_break_records, hard_break_type_counter = evaluate_hard_constraints(
            coords=coords,
            route_line_xy=route_line_xy,
            route_buffer_xy=route_buffer_xy,
            inv=inv,
            civil_hard_polys_xy=civil_hard_polys_xy,
            military_hard_polys_xy=military_hard_polys_xy,
            school_hard_union_xy=school_hard_union_xy,
        )
        hard_break_count = len(hard_break_records)

        weighted_score, score_detail = compute_weighted_score(
            hard_break_count=hard_break_count,
            avg_pop=avg_pop,
            w_hard=w_hard,
            w_pop=w_pop,
            hard_norm_cap=float(args.hard_norm_cap),
            pop_norm_cap=float(args.pop_norm_cap),
        )

        compliance = "不合规" if hard_break_count > 0 else "合规"
        risk_level = severity_from_metrics(hard_break_count, avg_pop)

        civil_nofly_features = xy_geoms_to_wgs_features(civil_hard_polys_xy, inv, "civil_airport_hard_nofly")
        military_nofly_features = xy_geoms_to_wgs_features(military_hard_polys_xy, inv, "military_hard_nofly")
        school_features = []
        if school_hard_union_xy is not None and (not school_hard_union_xy.is_empty):
            school_features = xy_geoms_to_wgs_features([school_hard_union_xy], inv, "school_kindergarten_hard")
        hv_line_features = xy_geoms_to_wgs_features(hv_power_lines_xy, inv, "high_voltage_power_line")

        hard_type_summary = "；".join([f"{k}:{v}" for k, v in sorted(hard_break_type_counter.items())])
        hard_detail_text = serialize_breach_details(hard_break_records)
        score_detail_text = (
            f"hard_norm={score_detail['hard_norm']:.6f}, "
            f"pop_norm={score_detail['pop_norm']:.6f}, "
            f"w_hard={score_detail['w_hard']:.4f}, "
            f"w_pop={score_detail['w_pop']:.4f}"
        )

        if compliance == "不合规" and args.replan_on_noncompliant != "off":
            replan_info["attempted"] = True
            should_run_replan = False
            if args.replan_on_noncompliant == "auto":
                should_run_replan = True
                replan_info["confirmed"] = True
            else:
                if sys.stdin is not None and sys.stdin.isatty():
                    prompt = (
                        f"[整改建议] 航线 {kml.name} 当前不合规（硬约束突破 {hard_break_count}）。"
                        "是否执行最小改动重规划？[y/N]: "
                    )
                    answer = input(prompt).strip().lower()
                    should_run_replan = answer in {"y", "yes", "1", "true", "是"}
                    replan_info["confirmed"] = bool(should_run_replan)
                    if not should_run_replan:
                        replan_info["status"] = "skipped"
                        replan_info["reason"] = "user_declined"
                else:
                    replan_info["status"] = "skipped"
                    replan_info["reason"] = "non_interactive_confirm_skipped"

            if should_run_replan:
                try:
                    policy = dict(REPLAN_POLICY_PRESETS.get(args.replan_policy, REPLAN_POLICY_PRESETS["balanced"]))
                    if args.replan_max_detour_ratio is not None:
                        policy["reference_max_detour_ratio"] = float(args.replan_max_detour_ratio)
                    if args.replan_max_mean_offset_m is not None:
                        policy["reference_max_mean_offset_m"] = float(args.replan_max_mean_offset_m)
                    if args.replan_reference_corridor_m is not None:
                        policy["reference_corridor_m"] = float(args.replan_reference_corridor_m)
                    if args.replan_reference_deviation_weight is not None:
                        policy["reference_deviation_weight"] = float(args.replan_reference_deviation_weight)
                    replan_info["policy"] = policy

                    primary_city = choose_primary_city(route_cities, coords)
                    if not primary_city:
                        raise RuntimeError("无法确定重规划城市")
                    replan_info["primary_city"] = primary_city
                    if len(route_cities) > 1:
                        replan_info["cross_city_warning"] = (
                            f"检测到跨城航线（{','.join(route_cities)}），重规划按主城市 {primary_city} 执行。"
                        )

                    if args.replan_output_dir:
                        replan_root = (root / args.replan_output_dir).resolve() / route_name
                    else:
                        replan_root = route_dir / "replan"
                    replan_root.mkdir(parents=True, exist_ok=True)
                    replan_name = f"{route_name}_replan"

                    replan_cmd = [
                        "python3",
                        str(root / "skills" / "plan-auto-route" / "scripts" / "plan_auto_route.py"),
                        "--city",
                        primary_city,
                        "--od-kml",
                        str(kml.resolve()),
                        "--name",
                        replan_name,
                        "--out-dir",
                        str(replan_root),
                        "--profile",
                        str(args.replan_profile),
                        "--select-candidate",
                        str(args.replan_select_candidate),
                        "--open-data-no-fly",
                        "--civil-airport-no-fly-geojson",
                        str(civil_airport_geojson),
                        "--reference-kml",
                        str(kml.resolve()),
                        "--reference-corridor-m",
                        str(policy["reference_corridor_m"]),
                        "--reference-deviation-weight",
                        str(policy["reference_deviation_weight"]),
                        "--reference-max-detour-ratio",
                        str(policy["reference_max_detour_ratio"]),
                        "--reference-max-mean-offset-m",
                        str(policy["reference_max_mean_offset_m"]),
                        "--no-write-snapshot",
                    ]
                    run_cmd(replan_cmd)

                    replan_kml = replan_root / f"{replan_name}.kml"
                    replan_html = replan_root / f"{replan_name}.html"
                    replan_meta = replan_root / f"{replan_name}_meta.json"
                    replan_candidates = replan_root / f"{replan_name}_candidates.json"
                    if not replan_kml.exists():
                        raise FileNotFoundError(f"replan kml missing: {replan_kml}")
                    if not replan_meta.exists():
                        raise FileNotFoundError(f"replan meta missing: {replan_meta}")

                    new_coords = parse_kml_coords(replan_kml)
                    old_len_m = route_length_m(coords)
                    new_len_m = route_length_m(new_coords)
                    off_mean, off_p90, off_max = route_offset_stats_m(coords, new_coords)

                    new_line_wgs, new_buffer_wgs = build_route_buffer(new_coords, BUFFER_EACH_SIDE_M)
                    new_points_wgs_lonlat = [(c[1], c[0]) for c in new_coords]
                    fwd_new, inv_new = plan_mod.build_projectors(new_points_wgs_lonlat)
                    new_line_xy = transform(fwd_new, new_line_wgs)
                    new_buffer_xy = transform(fwd_new, new_buffer_wgs)
                    new_start_xy = plan_mod.wgs_to_xy(fwd_new, new_points_wgs_lonlat[0][0], new_points_wgs_lonlat[0][1])
                    new_end_xy = plan_mod.wgs_to_xy(fwd_new, new_points_wgs_lonlat[-1][0], new_points_wgs_lonlat[-1][1])
                    new_bbox_wgs = plan_mod.route_bbox_wgs(new_points_wgs_lonlat, margin_m=float(args.route_bbox_margin_m))

                    (
                        _nofly_hard_new_xy,
                        _nofly_soft_new_xy,
                        _nofly_counter_new,
                        civil_hard_new_xy,
                        military_hard_new_xy,
                        _heli_soft_new_xy,
                    ) = plan_mod.fetch_open_data_no_fly_zones(
                        new_bbox_wgs,
                        fwd_new,
                        civil_airport_geojson=str(civil_airport_geojson),
                    )
                    school_hard_new_xy, _school_pts_new_xy, _school_counter_new, _school_items_new = plan_mod.fetch_school_kindergarten_zones(
                        new_bbox_wgs,
                        fwd_new,
                    )
                    school_hard_union_new_xy = unary_union(school_hard_new_xy) if school_hard_new_xy else None
                    if school_hard_union_new_xy is not None and (not school_hard_union_new_xy.is_empty):
                        relief_m_new = float(getattr(plan_mod, "SCHOOL_ENDPOINT_RELIEF_M", 55.0))
                        relief_new = Point(new_start_xy).buffer(relief_m_new).union(Point(new_end_xy).buffer(relief_m_new))
                        try:
                            school_hard_union_new_xy = school_hard_union_new_xy.difference(relief_new)
                        except Exception:
                            pass
                        if school_hard_union_new_xy.is_empty:
                            school_hard_union_new_xy = None

                    replan_hard_break_records, replan_hard_break_type_counter = evaluate_hard_constraints(
                        coords=new_coords,
                        route_line_xy=new_line_xy,
                        route_buffer_xy=new_buffer_xy,
                        inv=inv_new,
                        civil_hard_polys_xy=civil_hard_new_xy,
                        military_hard_polys_xy=military_hard_new_xy,
                        school_hard_union_xy=school_hard_union_new_xy,
                    )
                    replan_break_count = len(replan_hard_break_records)
                    replan_compliance = "不合规" if replan_break_count > 0 else "合规"
                    replan_status = "success" if replan_compliance == "合规" else "failed"

                    replan_info["status"] = replan_status
                    replan_info["reason"] = ""
                    replan_info["outputs"] = {
                        "kml": str(replan_kml.resolve()),
                        "html": str(replan_html.resolve()) if replan_html.exists() else "",
                        "meta_json": str(replan_meta.resolve()),
                        "candidates_json": str(replan_candidates.resolve()) if replan_candidates.exists() else "",
                    }
                    replan_info["comparison"] = {
                        "old_hard_break_count": int(hard_break_count),
                        "new_hard_break_count": int(replan_break_count),
                        "detour_ratio_vs_old": round(float(new_len_m) / max(1e-6, float(old_len_m)), 3),
                        "mean_offset_m": round(float(off_mean), 2),
                        "p90_offset_m": round(float(off_p90), 2),
                        "max_offset_m": round(float(off_max), 2),
                        "old_length_km": round(float(old_len_m) / 1000.0, 3),
                        "new_length_km": round(float(new_len_m) / 1000.0, 3),
                    }
                    replan_info["new_evaluation"] = {
                        "compliance": replan_compliance,
                        "hard_break_type_counter": replan_hard_break_type_counter,
                        "hard_break_points": replan_hard_break_records,
                    }
                    replan_coords_for_map = new_coords
                    replan_buffer_for_map = new_buffer_wgs
                    replan_hard_break_points_for_map = replan_hard_break_records
                except Exception as exc:
                    replan_info["status"] = "failed"
                    replan_info["reason"] = str(exc)

        route_html = route_dir / f"{route_name}_map.html"
        build_route_html(
            out_html=route_html,
            route_name=route_name,
            coords=coords,
            buffer_wgs=buffer_wgs,
            poi_features=poi_all,
            landuse_features=landuse_all,
            water_surface_features=water_surface_all,
            waterway_features=waterway_all,
            population_tile_dirs=pop_tile_dirs,
            road_features=roads_all,
            hsr_features=hsr_all,
            civil_nofly_features=civil_nofly_features,
            military_nofly_features=military_nofly_features,
            school_hard_features=school_features,
            hv_line_features=hv_line_features,
            hard_breach_points=hard_break_records,
            plan_mod=plan_mod,
            replan_coords=replan_coords_for_map,
            replan_buffer_wgs=replan_buffer_for_map,
            replan_hard_breach_points=replan_hard_break_points_for_map,
            replan_summary=replan_info,
        )

        meta_path = route_dir / f"{route_name}_meta_v3.json"
        meta_obj = {
            "workflow_version": "v3",
            "mode": "evaluate_existing_route_only",
            "route": {
                "name": kml.name,
                "path": str(kml.resolve()),
                "cities": route_cities,
                "max_true_height_m": round(float(max_alt), 2),
            },
            "constraints_standard": {
                "reference": "plan-auto-route:efficiency",
                "civil_nofly_rule": "civil airport polygon intersects route 100m buffer",
                "school_rule": "school/kindergarten hard zone intersects route centerline",
                "military_rule": "military hard no-fly intersects route centerline",
                "max_true_height_hard_limit_m": HARD_MAX_TRUE_HEIGHT_M,
            },
            "metrics": {
                "hard_constraint_break_count": int(hard_break_count),
                "avg_population_density_per_km2": round(float(avg_pop), 2),
                "normalized_weighted_score_0_100": float(weighted_score),
                "score_detail": score_detail,
                "compliance": compliance,
                "risk_level": risk_level,
            },
            "hard_constraint_break_type_counter": hard_break_type_counter,
            "hard_constraint_break_points": hard_break_records,
            "open_data_sources": {
                "city_data_cache": [str(city_cache_dir(root, city)) for city in route_cities],
                "civil_airport_dataset": str(civil_airport_geojson),
                "no_fly": nofly_counter,
                "school_kindergarten": school_counter,
                "line_risk": line_risk_counter_total,
                "obstacles": obstacle_counter,
                "soft_no_fly_zone_count": len(nofly_soft_polys_xy),
                "heli_soft_zone_count": len(heli_soft_polys_xy),
                "school_point_count": len(school_points_xy),
            },
            "outputs": {
                "map_html": str(route_html.resolve()),
                "meta_json": str(meta_path.resolve()),
            },
            "replan": replan_info,
            "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        meta_path.write_text(json.dumps(meta_obj, ensure_ascii=False, indent=2), encoding="utf-8")

        replan_status_text = "无需整改" if compliance == "合规" else "未执行"
        replan_suggestion_text = ""
        replan_kml_link = ""
        replan_meta_link = ""
        replan_compare_text = ""
        if compliance == "不合规":
            if args.replan_on_noncompliant == "off":
                replan_suggestion_text = "建议执行最小改动重规划"
            else:
                status = str(replan_info.get("status", ""))
                if status == "success":
                    replan_status_text = "整改成功"
                    new_comp = str(replan_info.get("new_evaluation", {}).get("compliance", ""))
                    replan_suggestion_text = f"整改后结论：{new_comp}"
                elif status == "failed":
                    replan_status_text = "整改失败"
                    reason = str(replan_info.get("reason", ""))
                    replan_suggestion_text = f"整改失败原因：{reason}" if reason else "整改失败，请人工复核"
                elif status == "skipped":
                    replan_status_text = "已跳过"
                    reason = str(replan_info.get("reason", ""))
                    replan_suggestion_text = "已跳过重规划" if not reason else f"已跳过重规划：{reason}"
                else:
                    replan_suggestion_text = "建议执行最小改动重规划"
        outputs = replan_info.get("outputs", {}) if isinstance(replan_info.get("outputs", {}), dict) else {}
        replan_kml_link = str(outputs.get("kml", ""))
        replan_meta_link = str(outputs.get("meta_json", ""))
        comp_obj = replan_info.get("comparison", {}) if isinstance(replan_info.get("comparison", {}), dict) else {}
        if comp_obj:
            replan_compare_text = (
                f"old_break={comp_obj.get('old_hard_break_count', '')}, "
                f"new_break={comp_obj.get('new_hard_break_count', '')}, "
                f"detour_vs_old={comp_obj.get('detour_ratio_vs_old', '')}, "
                f"mean_offset_m={comp_obj.get('mean_offset_m', '')}"
            )

        row = {
            "记录时间": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "航线名称": kml.name,
            "航线文件链接": str(kml.resolve()),
            "地图HTML链接": str(route_html.resolve()),
            "Meta链接": str(meta_path.resolve()),
            "覆盖城市": "、".join(route_cities),
            "航线最高高度(米)": round(float(max_alt), 2),
            "平均人口密度(人/平方公里)": round(float(avg_pop), 2),
            "硬约束突破次数": int(hard_break_count),
            "硬约束突破点": hard_detail_text,
            "硬约束类型统计": hard_type_summary,
            "归一化加权评分(0-100)": float(weighted_score),
            "合规结论": compliance,
            "风险等级": risk_level,
            "评分明细": score_detail_text,
            "整改状态": replan_status_text,
            "整改建议": replan_suggestion_text,
            "整改后航线链接": replan_kml_link,
            "整改后Meta链接": replan_meta_link,
            "整改对比摘要": replan_compare_text,
        }
        rows.append(row)

        summary_routes.append(
            {
                "route": kml.name,
                "cities": route_cities,
                "map_html": str(route_html.resolve()),
                "meta_json": str(meta_path.resolve()),
                "hard_constraint_break_count": int(hard_break_count),
                "avg_population_density_per_km2": round(float(avg_pop), 2),
                "weighted_score_0_100": float(weighted_score),
                "compliance": compliance,
                "risk_level": risk_level,
                "replan_status": replan_info.get("status", ""),
                "replan_outputs": replan_info.get("outputs", {}),
            }
        )

    df = pd.DataFrame(rows).reindex(columns=RA_COLUMNS)
    df.to_excel(xlsx_path, index=False)
    format_ra_xlsx(xlsx_path)

    summary_path = out_dir / "workflow_summary_v3.json"
    summary_path.write_text(
        json.dumps(
            {
                "workflow_version": "v3",
                "mode": "evaluate_existing_route_only",
                "routes": summary_routes,
                "cities": sorted(city_set),
                "ra_xlsx": str(xlsx_path.resolve()),
                "weights": {
                    "hard": round(w_hard, 6),
                    "population": round(w_pop, 6),
                    "hard_norm_cap": float(args.hard_norm_cap),
                    "population_norm_cap": float(args.pop_norm_cap),
                },
                "replan": {
                    "mode": args.replan_on_noncompliant,
                    "policy": args.replan_policy,
                },
                "civil_airport_dataset": str(civil_airport_geojson),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"航线风险评估 completed for {len(rows)} route(s)")
    print(f"RA: {xlsx_path}")
    print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
